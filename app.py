import streamlit as st
import pandas as pd
from io import BytesIO
import math
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

st.set_page_config(
    page_title="Email Marketing DB | Odoo",
    page_icon="📧",
    layout="wide"
)

st.title("📧 Generador de Base de Datos Email Marketing")
st.caption("Sube las bases exportadas desde Odoo para generar tu lista de email marketing.")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def read_file(file):
    name = file.name.lower()
    if name.endswith(".csv"):
        for enc in ["utf-8-sig", "utf-8", "latin-1"]:
            try:
                file.seek(0)
                return pd.read_csv(file, encoding=enc)
            except UnicodeDecodeError:
                continue
        raise ValueError("No se pudo leer el CSV con las codificaciones estándar.")
    return pd.read_excel(file)


def norm_key(val):
    if pd.isna(val) or str(val).strip() == "":
        return None
    return str(val).strip().lower()


def norm_email(val):
    if pd.isna(val) or str(val).strip() == "":
        return None
    return str(val).strip().lower()


def clean_str(val):
    if pd.isna(val):
        return ""
    return str(val).strip()


def validate_columns(df, required_cols, label):
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(
            f"**{label}**: faltan las columnas `{missing}`. "
            f"Columnas encontradas: `{df.columns.tolist()}`"
        )
        return False
    return True


def process_and_show(companies_keys, display_map, df_emp, df_cont):
    """Estadísticas + construcción del Excel dado un conjunto de empresas objetivo."""
    total_companies = len(companies_keys)

    # Filtrar empresas
    df_emp = df_emp.copy()
    df_emp["_key"] = df_emp["Nombre mostrado"].apply(norm_key)
    df_emp["_email"] = df_emp["Correo electrónico"].apply(norm_email)
    df_emp_f = df_emp[df_emp["_key"].isin(companies_keys)].copy()

    # Filtrar contactos usando claves canónicas de empresas
    canonical_keys = set(df_emp_f["_key"].dropna())
    df_cont = df_cont.copy()
    df_cont["_company_key"] = df_cont["Compañía relacionada"].apply(norm_key)
    df_cont["_email"] = df_cont["Correo electrónico"].apply(norm_email)
    df_cont_f = df_cont[df_cont["_company_key"].isin(canonical_keys)].copy()

    # Estadísticas
    comp_emp_email = set(df_emp_f.loc[df_emp_f["_email"].notna(), "_key"])
    comp_cont_email = set(df_cont_f.loc[df_cont_f["_email"].notna(), "_company_key"])
    comp_with_email = comp_emp_email | comp_cont_email
    total_with_email = len(comp_with_email)
    total_without_email = total_companies - total_with_email

    all_emails = (
        set(df_emp_f.loc[df_emp_f["_email"].notna(), "_email"])
        | set(df_cont_f.loc[df_cont_f["_email"].notna(), "_email"])
    )
    total_unique_emails = len(all_emails)

    # Construir Excel — contactos primero (prioridad en dedup)
    b_cont_valid = df_cont_f[df_cont_f["_email"].notna()].copy()
    cont_rows = pd.DataFrame({
        "name": b_cont_valid["Nombre"].apply(clean_str),
        "company name": b_cont_valid["Compañía relacionada"].apply(clean_str),
        "email": b_cont_valid["_email"],
    })

    b_emp_valid = df_emp_f[df_emp_f["_email"].notna()].copy()
    emp_rows = pd.DataFrame({
        "name": b_emp_valid["Nombre mostrado"].apply(clean_str),
        "company name": b_emp_valid["Nombre mostrado"].apply(clean_str),
        "email": b_emp_valid["_email"],
    })

    combined = pd.concat([cont_rows, emp_rows], ignore_index=True)
    combined = combined.drop_duplicates(subset=["email"], keep="first")
    combined = combined.sort_values(["company name", "name"]).reset_index(drop=True)

    # Mostrar resultados
    st.success("✅ Proceso completado")
    st.markdown("---")
    st.markdown("### 📊 Resumen")

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Empresas", total_companies)
    m2.metric("Con Correo", total_with_email)
    m3.metric("Sin Correo", total_without_email)
    m4.metric("Correos Únicos", total_unique_emails)

    found_in_emp = set(df_emp_f["_key"])
    missing = companies_keys - found_in_emp
    if missing:
        with st.expander(f"⚠️ {len(missing)} empresa(s) no encontradas en la base de empresas"):
            st.write(sorted(display_map.get(k, k) for k in missing))

    st.markdown("### 📋 Vista previa del Excel")
    st.dataframe(combined, use_container_width=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        combined.to_excel(writer, index=False, sheet_name="Email Marketing")
    output.seek(0)

    st.download_button(
        label="⬇️ Descargar Excel",
        data=output,
        file_name="email_marketing_db.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ─── Columnas esperadas ────────────────────────────────────────────────────────

PEDIDOS_COLS = ["Fecha creación", "Cliente"]
EMPRESAS_COLS = ["Nombre", "Nombre mostrado", "Correo electrónico"]
CONTACTOS_COLS = ["Nombre", "Compañía relacionada", "Correo electrónico"]
TALLER_EMPRESAS_COLS = ["Nombre", "Nombre mostrado", "Teléfono", "Móvil", "Correo electrónico"]
TALLER_CONTACTOS_COLS = ["Nombre", "Compañía relacionada", "Teléfono", "Móvil", "Correo electrónico"]


def process_taller(companies_keys, display_map, df_emp, df_cont):
    """Construye Excel con empresas y contactos internos incluyendo teléfono, móvil y correo."""
    # Filtrar empresas
    df_emp = df_emp.copy()
    df_emp["_key"] = df_emp["Nombre mostrado"].apply(norm_key)
    df_emp_f = df_emp[df_emp["_key"].isin(companies_keys)].copy()

    # Filtrar contactos usando las claves canónicas de las empresas encontradas
    canonical_keys = set(df_emp_f["_key"].dropna())
    df_cont = df_cont.copy()
    df_cont["_company_key"] = df_cont["Compañía relacionada"].apply(norm_key)
    df_cont_f = df_cont[df_cont["_company_key"].isin(canonical_keys)].copy()

    # Filas de empresas
    emp_rows = pd.DataFrame({
        "Nombre": df_emp_f["Nombre mostrado"].apply(clean_str),
        "Compañía relacionada": "",
        "Teléfono": df_emp_f["Teléfono"].apply(clean_str),
        "Móvil": df_emp_f["Móvil"].apply(clean_str),
        "Correo electrónico": df_emp_f["Correo electrónico"].apply(clean_str),
        "_sort_key": df_emp_f["Nombre mostrado"].apply(
            lambda v: str(v).strip().lower() if pd.notna(v) else ""
        ),
        "_is_company": True,
    })

    # Filas de contactos internos
    cont_rows = pd.DataFrame({
        "Nombre": df_cont_f["Nombre"].apply(clean_str),
        "Compañía relacionada": df_cont_f["Compañía relacionada"].apply(clean_str),
        "Teléfono": df_cont_f["Teléfono"].apply(clean_str),
        "Móvil": df_cont_f["Móvil"].apply(clean_str),
        "Correo electrónico": df_cont_f["Correo electrónico"].apply(clean_str),
        "_sort_key": df_cont_f["Compañía relacionada"].apply(
            lambda v: str(v).strip().lower() if pd.notna(v) else ""
        ),
        "_is_company": False,
    })

    combined = pd.concat([emp_rows, cont_rows], ignore_index=True)
    # Ordenar: primero por empresa (alfabético), luego empresa antes que sus contactos
    combined = combined.sort_values(
        ["_sort_key", "_is_company"], ascending=[True, False]
    ).reset_index(drop=True)
    combined = combined.drop(columns=["_is_company"])

    # Asignar grupo alternado por empresa (para colorear filas)
    group_idx = 0
    prev_key = None
    groups = []
    for key in combined["_sort_key"]:
        if key != prev_key:
            if prev_key is not None:
                group_idx += 1
            prev_key = key
        groups.append(group_idx % 2)
    combined["_group"] = groups
    combined = combined.drop(columns=["_sort_key"])

    # Estadísticas
    found_in_emp = set(df_emp_f["_key"])
    missing = companies_keys - found_in_emp

    st.success("✅ Proceso completado")
    st.markdown("---")
    st.markdown("### 📊 Resumen")

    m1, m2, m3 = st.columns(3)
    m1.metric("Empresas encontradas", len(df_emp_f))
    m2.metric("Contactos internos", len(df_cont_f))
    m3.metric("Total filas en Excel", len(combined))

    if missing:
        with st.expander(f"⚠️ {len(missing)} empresa(s) no encontradas en la base de empresas"):
            st.write(sorted(display_map.get(k, k) for k in missing))

    export_df = combined.drop(columns=["_group"])

    st.markdown("### 📋 Vista previa del Excel")
    st.dataframe(export_df, use_container_width=True)

    WHITE_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
    GRAY_FILL  = PatternFill(fill_type="solid", fgColor="D9D9D9")

    thin = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Anchos de columna (en unidades de carácter) — suman ~130, justo para A4 horizontal
    COL_WIDTHS = [35, 35, 15, 15, 30]
    LINE_HEIGHT_PT = 15  # puntos por línea de texto

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Taller")
        ws = writer.sheets["Taller"]

        # Anchos de columna
        for col_idx, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # Bordes, ajuste de texto y altura de fila automática
        for row_idx in range(1, len(export_df) + 2):
            max_lines = 1
            for col_idx, col_w in enumerate(COL_WIDTHS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if row_idx == 1:
                    cell.font = Font(bold=True)
                if row_idx >= 2:
                    text = str(cell.value) if cell.value else ""
                    lines = max(1, math.ceil(len(text) / col_w)) if text else 1
                    max_lines = max(max_lines, lines)
            if row_idx >= 2:
                ws.row_dimensions[row_idx].height = max_lines * LINE_HEIGHT_PT

        # Colores de fila alternados por empresa
        for row_idx, group in enumerate(combined["_group"], start=2):
            fill = WHITE_FILL if group == 0 else GRAY_FILL
            for col_idx in range(1, len(export_df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        # Configuración de página para exportar PDF en A4 horizontal ajustado al ancho
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9  # A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    output.seek(0)

    st.download_button(
        label="⬇️ Descargar Excel",
        data=output,
        file_name="taller_contactos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ─── Pestañas ─────────────────────────────────────────────────────────────────

tab1, tab2, tab3 = st.tabs(["🛒 Compraron", "📋 Cotizaron", "🔧 Taller"])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — COMPRARON
# ══════════════════════════════════════════════════════════════════════════════

with tab1:
    st.subheader("1. Base de Pedidos Odoo")
    st.caption("Filtro en Odoo: part number de la marca + estado = pedido de venta")
    st.caption("Columnas requeridas: `Fecha creación`, `Cliente`")
    t1_file1 = st.file_uploader("Subir Base 1", type=["xlsx", "csv"], key="t1_b1")

    st.divider()

    st.subheader("2. Base de Empresas")
    st.caption("Filtro en Odoo: Compañía relacionada - No está establecida")
    st.caption("Columnas requeridas: `Nombre`, `Nombre mostrado`, `Correo electrónico`")
    t1_file2 = st.file_uploader("Subir Base 2", type=["xlsx", "csv"], key="t1_b2")

    st.divider()

    st.subheader("3. Base de Contactos Internos")
    st.caption("Filtro en Odoo: Compañía relacionada - Está establecida")
    st.caption("Columnas requeridas: `Nombre`, `Compañía relacionada`, `Correo electrónico`")
    t1_file3 = st.file_uploader("Subir Base 3", type=["xlsx", "csv"], key="t1_b3")

    st.divider()

    t1_df1 = t1_df2 = t1_df3 = None
    t1_ok1 = t1_ok2 = t1_ok3 = False

    if t1_file1:
        try:
            t1_df1 = read_file(t1_file1)
            t1_ok1 = validate_columns(t1_df1, PEDIDOS_COLS, "Base 1")
        except Exception as e:
            st.error(f"Error al leer Base 1: {e}")

    if t1_file2:
        try:
            t1_df2 = read_file(t1_file2)
            t1_ok2 = validate_columns(t1_df2, EMPRESAS_COLS, "Base 2")
        except Exception as e:
            st.error(f"Error al leer Base 2: {e}")

    if t1_file3:
        try:
            t1_df3 = read_file(t1_file3)
            t1_ok3 = validate_columns(t1_df3, CONTACTOS_COLS, "Base 3")
        except Exception as e:
            st.error(f"Error al leer Base 3: {e}")

    if not (t1_ok1 and t1_ok2 and t1_ok3):
        st.info("Sube las 3 bases con las columnas correctas para habilitar el procesamiento.")
    else:
        if st.button("🚀 Procesar", type="primary", use_container_width=True, key="t1_run"):
            with st.spinner("Procesando..."):
                companies_keys = {
                    k for k in (norm_key(v) for v in t1_df1["Cliente"].dropna())
                    if k is not None
                }
                display_map = {}
                for v in t1_df1["Cliente"].dropna():
                    k = norm_key(v)
                    if k and k not in display_map:
                        display_map[k] = clean_str(v)

                process_and_show(companies_keys, display_map, t1_df2, t1_df3)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — COTIZARON (Y NO COMPRARON)
# ══════════════════════════════════════════════════════════════════════════════

with tab2:
    st.subheader("1. Base de Pedidos Odoo (Compras)")
    st.caption("Filtro en Odoo: part number de la marca + estado = pedido de venta")
    st.caption("Columnas requeridas: `Fecha creación`, `Cliente`")
    t2_file1 = st.file_uploader("Subir Base 1", type=["xlsx", "csv"], key="t2_b1")

    st.divider()

    st.subheader("2. Base de Cotizaciones Odoo")
    st.caption("Filtro en Odoo: part number de la marca")
    st.caption("Columnas requeridas: `Fecha creación`, `Cliente`")
    t2_file2 = st.file_uploader("Subir Base 2", type=["xlsx", "csv"], key="t2_b2")

    st.divider()

    st.subheader("3. Base de Empresas")
    st.caption("Filtro en Odoo: Compañía relacionada - No está establecida")
    st.caption("Columnas requeridas: `Nombre`, `Nombre mostrado`, `Correo electrónico`")
    t2_file3 = st.file_uploader("Subir Base 3", type=["xlsx", "csv"], key="t2_b3")

    st.divider()

    st.subheader("4. Base de Contactos Internos")
    st.caption("Filtro en Odoo: Compañía relacionada - Está establecida")
    st.caption("Columnas requeridas: `Nombre`, `Compañía relacionada`, `Correo electrónico`")
    t2_file4 = st.file_uploader("Subir Base 4", type=["xlsx", "csv"], key="t2_b4")

    st.divider()

    t2_df1 = t2_df2 = t2_df3 = t2_df4 = None
    t2_ok1 = t2_ok2 = t2_ok3 = t2_ok4 = False

    if t2_file1:
        try:
            t2_df1 = read_file(t2_file1)
            t2_ok1 = validate_columns(t2_df1, PEDIDOS_COLS, "Base 1 (Compras)")
        except Exception as e:
            st.error(f"Error al leer Base 1: {e}")

    if t2_file2:
        try:
            t2_df2 = read_file(t2_file2)
            t2_ok2 = validate_columns(t2_df2, PEDIDOS_COLS, "Base 2 (Cotizaciones)")
        except Exception as e:
            st.error(f"Error al leer Base 2: {e}")

    if t2_file3:
        try:
            t2_df3 = read_file(t2_file3)
            t2_ok3 = validate_columns(t2_df3, EMPRESAS_COLS, "Base 3 (Empresas)")
        except Exception as e:
            st.error(f"Error al leer Base 3: {e}")

    if t2_file4:
        try:
            t2_df4 = read_file(t2_file4)
            t2_ok4 = validate_columns(t2_df4, CONTACTOS_COLS, "Base 4 (Contactos)")
        except Exception as e:
            st.error(f"Error al leer Base 4: {e}")

    if not (t2_ok1 and t2_ok2 and t2_ok3 and t2_ok4):
        st.info("Sube las 4 bases con las columnas correctas para habilitar el procesamiento.")
    else:
        if st.button("🚀 Procesar", type="primary", use_container_width=True, key="t2_run"):
            with st.spinner("Procesando..."):
                buyers = {
                    k for k in (norm_key(v) for v in t2_df1["Cliente"].dropna())
                    if k is not None
                }
                quoters_all = {
                    k for k in (norm_key(v) for v in t2_df2["Cliente"].dropna())
                    if k is not None
                }
                # Solo quienes cotizaron pero NO compraron
                companies_keys = quoters_all - buyers

                display_map = {}
                for v in t2_df2["Cliente"].dropna():
                    k = norm_key(v)
                    if k and k not in display_map:
                        display_map[k] = clean_str(v)

                if not companies_keys:
                    st.warning("No hay empresas que hayan cotizado sin haber comprado.")
                else:
                    st.info(
                        f"Cotizaciones totales: **{len(quoters_all)}** empresas · "
                        f"Compradores excluidos: **{len(buyers & quoters_all)}** · "
                        f"Resultado: **{len(companies_keys)}** empresas"
                    )
                    process_and_show(companies_keys, display_map, t2_df3, t2_df4)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — TALLER
# ══════════════════════════════════════════════════════════════════════════════

with tab3:
    st.subheader("1. Base de Compras de la Marca")
    st.caption("Filtro en Odoo: part number de la marca + estado = pedido de venta")
    st.caption("Columnas requeridas: `Fecha creación`, `Cliente`")
    t3_file1 = st.file_uploader("Subir Base 1", type=["xlsx", "csv"], key="t3_b1")

    st.divider()

    st.subheader("2. Base de Empresas")
    st.caption("Filtro en Odoo: Compañía relacionada - No está establecida")
    st.caption("Columnas requeridas: `Nombre`, `Nombre mostrado`, `Teléfono`, `Móvil`, `Correo electrónico`")
    t3_file2 = st.file_uploader("Subir Base 2", type=["xlsx", "csv"], key="t3_b2")

    st.divider()

    st.subheader("3. Base de Contactos Internos")
    st.caption("Filtro en Odoo: Compañía relacionada - Está establecida")
    st.caption("Columnas requeridas: `Nombre`, `Compañía relacionada`, `Teléfono`, `Móvil`, `Correo electrónico`")
    t3_file3 = st.file_uploader("Subir Base 3", type=["xlsx", "csv"], key="t3_b3")

    st.divider()

    t3_df1 = t3_df2 = t3_df3 = None
    t3_ok1 = t3_ok2 = t3_ok3 = False

    if t3_file1:
        try:
            t3_df1 = read_file(t3_file1)
            t3_ok1 = validate_columns(t3_df1, PEDIDOS_COLS, "Base 1")
        except Exception as e:
            st.error(f"Error al leer Base 1: {e}")

    if t3_file2:
        try:
            t3_df2 = read_file(t3_file2)
            t3_ok2 = validate_columns(t3_df2, TALLER_EMPRESAS_COLS, "Base 2")
        except Exception as e:
            st.error(f"Error al leer Base 2: {e}")

    if t3_file3:
        try:
            t3_df3 = read_file(t3_file3)
            t3_ok3 = validate_columns(t3_df3, TALLER_CONTACTOS_COLS, "Base 3")
        except Exception as e:
            st.error(f"Error al leer Base 3: {e}")

    if not (t3_ok1 and t3_ok2 and t3_ok3):
        st.info("Sube las 3 bases con las columnas correctas para habilitar el procesamiento.")
    else:
        if st.button("🚀 Procesar", type="primary", use_container_width=True, key="t3_run"):
            with st.spinner("Procesando..."):
                companies_keys = {
                    k for k in (norm_key(v) for v in t3_df1["Cliente"].dropna())
                    if k is not None
                }
                display_map = {}
                for v in t3_df1["Cliente"].dropna():
                    k = norm_key(v)
                    if k and k not in display_map:
                        display_map[k] = clean_str(v)

                process_taller(companies_keys, display_map, t3_df2, t3_df3)
